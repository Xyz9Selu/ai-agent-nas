import io
import json
import logging
import os
from datetime import datetime
from typing import Any, Iterable, Iterator

import openpyxl

import db
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload


SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/spreadsheets.readonly",
]

BATCH_SIZE = int(os.getenv("SAP_PARSER_BATCH_SIZE", "5000"))
DB_INSERT_BATCH_SIZE = int(os.getenv("SAP_PARSER_DB_BATCH_SIZE", "500"))

logger = logging.getLogger(__name__)


def build_drive_service(access_token: str):
    creds = Credentials(token=access_token, scopes=SCOPES)
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def build_sheets_service(access_token: str):
    creds = Credentials(token=access_token, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds, cache_discovery=False)


def get_drive_file_metadata(service, file_id: str) -> dict[str, Any]:
    return (
        service.files()
        .get(fileId=file_id, fields="mimeType,name")
        .execute()
    )


def download_file_bytes(service, file_id: str) -> bytes:
    """Download a Drive file as raw bytes (for XLSX or other binary types)."""
    request = service.files().get_media(fileId=file_id)
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _status, done = downloader.next_chunk()
    fh.seek(0)
    return fh.read()


def _col_index_to_letter(idx: int) -> str:
    """Convert 1-based column index to Excel column letter (1 -> A, 27 -> AA)."""
    result = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def get_google_sheet_chosen_title(
    sheets_service,
    spreadsheet_id: str,
    *,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
) -> str:
    """Return the title of the chosen sheet (by sheet_id or sheet_name or first sheet)."""
    meta = (
        sheets_service.spreadsheets()
        .get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title))",
        )
        .execute()
    )
    sheets = meta.get("sheets", [])
    if not sheets:
        raise ValueError("Spreadsheet has no sheets")
    if sheet_id is not None:
        for s in sheets:
            if s["properties"].get("sheetId") == sheet_id:
                return s["properties"]["title"]
        raise ValueError(f"Sheet not found: sheet_id={sheet_id}")
    if sheet_name is not None:
        for s in sheets:
            if s["properties"].get("title") == sheet_name:
                return s["properties"]["title"]
        raise ValueError(f"Sheet not found: sheet_name={repr(sheet_name)}")
    return sheets[0]["properties"]["title"]


def iter_google_sheet_rows(
    sheets_service,
    spreadsheet_id: str,
    batch_size: int = BATCH_SIZE,
    *,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
) -> Iterator[list[str]]:
    """Yield all rows from a sheet of a Google Spreadsheet.

    If sheet_id is given, use that sheet (takes precedence over sheet_name).
    Else if sheet_name is given, use the sheet with that title.
    Else use the first sheet. Raises ValueError if a specified sheet is not found.
    """
    meta = (
        sheets_service.spreadsheets()
        .get(
            spreadsheetId=spreadsheet_id,
            fields="sheets(properties(sheetId,title,gridProperties(rowCount,columnCount)))",
        )
        .execute()
    )
    sheets = meta.get("sheets", [])
    if not sheets:
        return

    chosen = None
    if sheet_id is not None:
        for s in sheets:
            if s["properties"].get("sheetId") == sheet_id:
                chosen = s
                break
        if chosen is None:
            raise ValueError(f"Sheet not found: sheet_id={sheet_id}")
    elif sheet_name is not None:
        for s in sheets:
            if s["properties"].get("title") == sheet_name:
                chosen = s
                break
        if chosen is None:
            raise ValueError(f"Sheet not found: sheet_name={repr(sheet_name)}")
    else:
        chosen = sheets[0]

    props = chosen["properties"]
    title = props["title"]
    grid = props.get("gridProperties", {})
    row_count = int(grid.get("rowCount", 0) or 0)
    col_count = int(grid.get("columnCount", 0) or 0)

    if row_count == 0 or col_count == 0:
        return

    end_col_letter = _col_index_to_letter(col_count)

    for start_row in range(1, row_count + 1, batch_size):
        end_row = min(row_count, start_row + batch_size - 1)
        range_str = f"'{title}'!A{start_row}:{end_col_letter}{end_row}"
        resp = (
            sheets_service.spreadsheets()
            .values()
            .get(spreadsheetId=spreadsheet_id, range=range_str)
            .execute()
        )
        values: list[list[Any]] = resp.get("values", [])
        if not values:
            continue

        for row in values:
            # Normalize row length and cast to str.
            normalized_row = [
                "" if i >= len(row) or row[i] is None else str(row[i])
                for i in range(col_count)
            ]
            yield normalized_row


def iter_xlsx_rows(file_bytes: bytes) -> Iterator[list[str]]:
    """Yield rows from an XLSX file using openpyxl read_only mode."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        ws = wb.worksheets[0]
        for row in ws.iter_rows(values_only=True):
            if row is None:
                continue
            # Convert to list of strings, preserving empty cells as "".
            normalized_row = [
                "" if cell is None else str(cell)
                for cell in row
            ]
            yield normalized_row
    finally:
        wb.close()


def get_xlsx_sheet_title_and_rows(
    file_bytes: bytes,
) -> tuple[str, Iterator[list[str]]]:
    """Return (first sheet title, row iterator) for an XLSX file."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    title = ws.title

    def rows() -> Iterator[list[str]]:
        try:
            for row in ws.iter_rows(values_only=True):
                if row is None:
                    continue
                yield [
                    "" if cell is None else str(cell)
                    for cell in row
                ]
        finally:
            wb.close()

    return (title, rows())


def _is_separator_line(line: str) -> bool:
    """True if line contains only |, -, and spaces."""
    s = line.strip()
    return all(c in "|- \t" for c in s) and len(s) > 0


def parse_txt_sap_report(content: str) -> tuple[list[str], Iterator[list[str]]]:
    """Parse SAP-style pipe-delimited report. Returns (column_names, row_iterator).

    Column names are de-duplicated (suffix _2, _3, ...). Rows are lists of strings
    with same length as column_names.
    """
    lines = content.splitlines()
    header_row: list[str] | None = None
    header_idx = -1

    for i, line in enumerate(lines):
        if not line.strip().startswith("|"):
            continue
        parts = [p.strip() for p in line.split("|")]
        # Drop leading/trailing empty from split on |
        if parts and parts[0] == "":
            parts = parts[1:]
        if parts and parts[-1] == "":
            parts = parts[:-1]
        if len(parts) < 2:
            continue
        if _is_separator_line(line):
            continue
        # Treat as header: has multiple segments that are not all dashes
        if not all(segment.replace("-", "").replace(" ", "") == "" for segment in parts):
            header_row = parts
            header_idx = i
            break

    if not header_row:
        return ([], iter([]))

    # De-duplicate column names
    seen: dict[str, int] = {}
    column_names: list[str] = []
    for h in header_row:
        name = (h or "").strip() or "col"
        if name in seen:
            seen[name] += 1
            column_names.append(f"{name}_{seen[name]}")
        else:
            seen[name] = 0
            column_names.append(name)

    def row_iter() -> Iterator[list[str]]:
        ncols = len(column_names)
        for i in range(header_idx + 1, len(lines)):
            line = lines[i]
            if not line.strip().startswith("|"):
                continue
            if _is_separator_line(line):
                continue
            parts = [p.strip() for p in line.split("|")]
            if parts and parts[0] == "":
                parts = parts[1:]
            if parts and parts[-1] == "":
                parts = parts[:-1]
            # Normalize length
            row = [parts[j] if j < len(parts) else "" for j in range(ncols)]
            yield row

    return (column_names, row_iter())


def _header_and_records_from_row_iter(
    rows: Iterator[list[str]],
) -> tuple[list[str] | None, Iterator[dict[str, Any]]]:
    """Detect header (first row with >5 non-empty cells), de-dup names, then yield records.
    Returns (final_header, record_iterator). If no header found, (None, empty iterator).
    """
    header: list[str] | None = None
    valid_col_idx: list[int] = []
    final_header: list[str] = []
    first_col_idx: int | None = None
    first_header_value: str | None = None

    for row in rows:
        actual_content = [c for c in row if str(c).strip() != ""]
        if len(actual_content) > 5:
            header = row
            valid_col_idx = [i for i, h in enumerate(header) if str(h).strip() != ""]
            filtered_header = [str(header[i]).strip() for i in valid_col_idx]
            if not valid_col_idx:
                return (None, iter([]))
            first_col_idx = valid_col_idx[0]
            first_header_value = filtered_header[0] if filtered_header else ""
            seen: dict[str, int] = {}
            for h in filtered_header:
                if h in seen:
                    seen[h] += 1
                    final_header.append(f"{h}_{seen[h]}")
                else:
                    seen[h] = 0
                    final_header.append(h)
            break

    if not final_header:
        return (None, iter([]))

    def record_iter() -> Iterator[dict[str, Any]]:
        for row in rows:
            if not any(str(c).strip() for c in row):
                continue
            if (
                first_col_idx is not None
                and first_header_value is not None
                and first_col_idx < len(row)
                and str(row[first_col_idx]) == first_header_value
            ):
                continue
            record = {
                col_name: (row[col_idx] if col_idx < len(row) else "")
                for col_idx, col_name in zip(valid_col_idx, final_header)
            }
            yield record

    return (final_header, record_iter())


def _stream_clean_and_write(
    rows: Iterable[list[str]],
    *,
    output_path: str | None = None,
    dataset_id: str | None = None,
    db_conn: Any = None,
) -> tuple[int, int]:
    """Apply clean_and_process semantics in a streaming way. Write to JSONL and/or DB.

    Args:
        rows: Raw row data.
        output_path: If set, write records to JSONL file.
        dataset_id: If set with db_conn, write to DB.
        db_conn: If set with dataset_id, write to DB.

    Returns:
        Tuple of (total_rows, rows_inserted).
    """
    header: list[str] | None = None
    valid_col_idx: list[int] = []
    final_header: list[str] = []
    first_col_idx: int | None = None
    first_header_value: str | None = None
    total_rows = 0
    rows_inserted = 0
    db_batch: list[dict[str, Any]] = []

    file_handle = (
        open(output_path, "w", encoding="utf-8") if output_path else None
    )
    try:
        for row in rows:
            # Find header row: first row with more than 5 non-empty cells.
            if header is None:
                actual_content = [cell for cell in row if str(cell).strip() != ""]
                if len(actual_content) > 5:
                    header = row
                    valid_col_idx = [
                        i for i, h in enumerate(header) if str(h).strip() != ""
                    ]
                    filtered_header = [
                        str(header[i]).strip() for i in valid_col_idx
                    ]

                    if not valid_col_idx:
                        # No useful header, skip entire stream.
                        return (0, 0)

                    first_col_idx = valid_col_idx[0]
                    first_header_value = filtered_header[0] if filtered_header else ""

                    # De-duplicate column names.
                    seen: dict[str, int] = {}
                    final_header = []
                    for h in filtered_header:
                        if h in seen:
                            seen[h] += 1
                            final_header.append(f"{h}_{seen[h]}")
                        else:
                            seen[h] = 0
                            final_header.append(h)
                # Still searching for header, skip this row for data purposes.
                continue

            # After header is found: apply same filters as clean_and_process.
            if not any(str(cell).strip() for cell in row):
                continue

            if (
                first_col_idx is not None
                and first_header_value is not None
                and first_col_idx < len(row)
                and str(row[first_col_idx]) == first_header_value
            ):
                # Skip repeated header rows.
                continue

            record: dict[str, Any] = {}
            for col_idx, col_name in zip(valid_col_idx, final_header):
                value = row[col_idx] if col_idx < len(row) else ""
                record[col_name] = value

            if file_handle:
                file_handle.write(json.dumps(record, ensure_ascii=False))
                file_handle.write("\n")
            total_rows += 1

            # Accumulate for DB batch insert.
            if db_conn and dataset_id:
                db_batch.append(record)
                if len(db_batch) >= DB_INSERT_BATCH_SIZE:
                    try:
                        rows_inserted += db.insert_rows_batch(
                            db_conn, dataset_id, db_batch
                        )
                    except Exception as e:
                        logger.error("DB batch insert failed: %s", e)
                        raise
                    db_batch = []

            if total_rows % BATCH_SIZE == 0:
                logger.info("[sap_parser] processed %d rows so far.", total_rows)

        # Flush remaining DB batch.
        if db_conn and dataset_id and db_batch:
            try:
                rows_inserted += db.insert_rows_batch(db_conn, dataset_id, db_batch)
            except Exception as e:
                logger.error("DB batch insert failed: %s", e)
                raise
    finally:
        if file_handle:
            file_handle.close()

    return (total_rows, rows_inserted)


def write_sap_sheet_to_database(
    file_id: str,
    access_token: str,
    *,
    dataset_id: str | None = None,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Parse SAP sheet from Google Drive and write rows to database."""
    drive_service = build_drive_service(access_token)
    metadata = get_drive_file_metadata(drive_service, file_id)
    mime_type = metadata.get("mimeType", "")
    name = metadata.get("name", "")

    effective_dataset_id = dataset_id if dataset_id is not None else file_id

    # Google Sheet: use Sheets API to read in batches.
    if mime_type == "application/vnd.google-apps.spreadsheet":
        sheets_service = build_sheets_service(access_token)
        rows_iter = iter_google_sheet_rows(
            sheets_service,
            spreadsheet_id=file_id,
            batch_size=BATCH_SIZE,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
        )
    else:
        # XLSX or other tabular file: download and stream via openpyxl.
        file_bytes = download_file_bytes(drive_service, file_id)
        rows_iter = iter_xlsx_rows(file_bytes)

    conn = db.get_db_connection()
    if conn is None:
        raise RuntimeError(
            "Database connection failed: missing POSTGRES_HOST, POSTGRES_DB, "
            "POSTGRES_USER, or POSTGRES_PASSWORD"
        )
    db.ensure_document_rows_table(conn)

    try:
        total_rows, rows_inserted = _stream_clean_and_write(
            rows_iter,
            dataset_id=effective_dataset_id,
            db_conn=conn,
        )
    finally:
        conn.close()

    result: dict[str, Any] = {
        "file_id": file_id,
        "name": name,
        "mime_type": mime_type,
        "total_rows": total_rows,
        "rows_inserted": rows_inserted,
    }
    return result


def write_sap_sheet_to_file(
    file_id: str,
    access_token: str,
    *,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Parse SAP sheet from Google Drive and write rows to JSONL file."""
    drive_service = build_drive_service(access_token)
    metadata = get_drive_file_metadata(drive_service, file_id)
    mime_type = metadata.get("mimeType", "")
    name = metadata.get("name", "")

    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    output_filename = f"{file_id}_{timestamp}.jsonl"
    output_path = os.path.join(os.getcwd(), output_filename)

    if mime_type == "application/vnd.google-apps.spreadsheet":
        sheets_service = build_sheets_service(access_token)
        rows_iter = iter_google_sheet_rows(
            sheets_service,
            spreadsheet_id=file_id,
            batch_size=BATCH_SIZE,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
        )
    else:
        file_bytes = download_file_bytes(drive_service, file_id)
        rows_iter = iter_xlsx_rows(file_bytes)

    total_rows, _ = _stream_clean_and_write(
        rows_iter,
        output_path=output_path,
    )

    return {
        "file_id": file_id,
        "name": name,
        "mime_type": mime_type,
        "output_file": output_path,
        "total_rows": total_rows,
    }


def write_sap_sheet_to_table(
    file_id: str,
    access_token: str,
    *,
    sheet_id: int | None = None,
    sheet_name: str | None = None,
    table_prefix: str | None = None,
) -> dict[str, Any]:
    """Parse SAP sheet or TXT report from Google Drive and write to a dedicated table.
    Table name = optional table_prefix + sanitized(document_name + sheet_name). Schema from source header.
    If table_prefix is not provided, the table name has no prefix.
    Returns table_name, schema, total_rows, rows_inserted, file_id, name, mime_type.
    """
    drive_service = build_drive_service(access_token)
    metadata = get_drive_file_metadata(drive_service, file_id)
    mime_type = metadata.get("mimeType", "")
    document_name = metadata.get("name", "")

    column_names: list[str]
    record_iter: Iterator[dict[str, Any]]
    sheet_title: str

    if mime_type == "application/vnd.google-apps.spreadsheet":
        sheets_service = build_sheets_service(access_token)
        sheet_title = get_google_sheet_chosen_title(
            sheets_service,
            file_id,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
        )
        rows_iter = iter_google_sheet_rows(
            sheets_service,
            spreadsheet_id=file_id,
            batch_size=BATCH_SIZE,
            sheet_id=sheet_id,
            sheet_name=sheet_name,
        )
        header, record_iter = _header_and_records_from_row_iter(rows_iter)
        if header is None:
            raise ValueError("No header row found in sheet")
        column_names = header
    elif mime_type in ("text/plain", "application/octet-stream"):
        file_bytes = download_file_bytes(drive_service, file_id)
        try:
            content = file_bytes.decode("utf-8")
        except UnicodeDecodeError as e:
            raise ValueError(f"File is not valid UTF-8: {e}") from e
        column_names, row_iter = parse_txt_sap_report(content)
        if not column_names:
            raise ValueError("No header row found in TXT report")

        def records() -> Iterator[dict[str, Any]]:
            for row in row_iter:
                yield dict(zip(column_names, row))

        record_iter = records()
        sheet_title = "data"
    else:
        # XLSX or other binary: download and use openpyxl
        file_bytes = download_file_bytes(drive_service, file_id)
        sheet_title, rows_iter = get_xlsx_sheet_title_and_rows(file_bytes)
        header, record_iter = _header_and_records_from_row_iter(rows_iter)
        if header is None:
            raise ValueError("No header row found in spreadsheet")
        column_names = header

    table_name = db.sanitize_table_name(document_name, sheet_title, prefix=table_prefix)

    conn = db.get_db_connection()
    if conn is None:
        raise RuntimeError(
            "Database connection failed: missing POSTGRES_HOST, POSTGRES_DB, "
            "POSTGRES_USER, or POSTGRES_PASSWORD"
        )
    try:
        total_rows, rows_inserted = db.ensure_import_table_and_insert(
            conn,
            table_name,
            column_names,
            record_iter,
            batch_size=DB_INSERT_BATCH_SIZE,
        )
    finally:
        conn.close()

    # Schema uses actual table column names (sanitized), not original headers
    sanitized_columns = db.sanitize_column_names(column_names)
    schema = [{"name": c, "type": "text"} for c in sanitized_columns]
    return {
        "table_name": table_name,
        "schema": schema,
        "total_rows": total_rows,
        "rows_inserted": rows_inserted,
        "file_id": file_id,
        "name": document_name,
        "mime_type": mime_type,
    }

